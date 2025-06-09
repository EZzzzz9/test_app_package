import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import load_workbook
import random
import csv
import os
import json

class TestApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Тестовое приложение")
        self.root.geometry("800x600")
        self.questions = []
        self.current_question = 0
        self.correct = 0
        self.incorrect = 0
        self.wrong_answers = []
        self.selected_file = "test_data.xlsx"
        self.stats_file = "statistics.json"
        self.stats = self.load_stats()
        self.main_menu()

    def main_menu(self):
        self.clear_window()
        tk.Label(self.root, text="Главное меню", font=("Arial", 24)).pack(pady=20)

        tk.Button(self.root, text="Начать тест", command=self.choose_test_mode, width=25, height=2).pack(pady=10)
        tk.Button(self.root, text="Загрузить ошибки из CSV", command=self.load_errors_csv, width=25, height=2).pack(pady=10)
        tk.Button(self.root, text="Статистика", command=self.show_stats, width=25, height=2).pack(pady=10)
        tk.Button(self.root, text="Выход", command=self.root.quit, width=25, height=2).pack(pady=10)

    def show_question(self):
        self.clear_window()
        q = self.test_questions[self.current_question]
        tk.Label(self.root, text=f"Вопрос {self.current_question + 1} из {len(self.test_questions)}", font=("Arial", 14)).pack(pady=10)
        tk.Label(self.root, text=q["question"], wraplength=750, font=("Arial", 16), justify="left").pack(pady=20)

        self.buttons = []
        for i, opt in enumerate(q["options"]):
            if opt.strip() == "":
                continue
            btn = tk.Button(self.root,
                            text=f"{chr(65 + i)}: {opt}",
                            wraplength=500,
                            justify="left",
                            anchor="w",
                            width=70,
                            command=lambda i=i: self.check_answer(i))
            btn.pack(pady=5)
            self.buttons.append(btn)

    def check_answer(self, selected_index):
        q = self.test_questions[self.current_question]
        correct_index = ord(q["answer"]) - 65 if q["answer"] else -1

        for i, btn in enumerate(self.buttons):
            if i == correct_index:
                btn.config(bg="green", fg="white")
            elif i == selected_index:
                btn.config(bg="red", fg="white")
            btn.config(state="disabled")

        if selected_index == correct_index:
            self.correct += 1
        else:
            self.incorrect += 1
            if q not in self.wrong_answers:
                self.wrong_answers.append(q)

        self.root.after(1000, self.next_question)

    def next_question(self):
        self.current_question += 1
        if self.current_question < len(self.test_questions):
            self.show_question()
        else:
            self.show_result()

    def show_result(self):
        self.clear_window()
        tk.Label(self.root, text="Тест завершён!", font=("Arial", 24)).pack(pady=20)
        tk.Label(self.root, text=f"Правильных ответов: {self.correct}", font=("Arial", 16)).pack(pady=5)
        tk.Label(self.root, text=f"Неправильных ответов: {self.incorrect}", font=("Arial", 16)).pack(pady=5)

        self.save_stats(self.correct, self.incorrect)

        if self.incorrect > 0:
            tk.Button(self.root, text="Прорешать ошибки", command=self.retry_wrong, width=25, height=2).pack(pady=10)
            tk.Button(self.root, text="Сохранить ошибки в CSV", command=self.save_errors, width=30, height=2).pack(pady=5)
        else:
            tk.Label(self.root, text="Все ответы верны!").pack(pady=10)
            tk.Button(self.root, text="Пройти тест заново", command=self.choose_test_mode, width=25, height=2).pack(pady=10)

        tk.Button(self.root, text="В меню", command=self.main_menu, width=25, height=2).pack(pady=20)

    def retry_wrong(self):
        if not self.wrong_answers:
            messagebox.showinfo("Ошибок нет", "Ошибок для прорешивания нет.")
            return
        self.test_questions = self.wrong_answers.copy()
        self.current_question = 0
        self.correct = 0
        self.incorrect = 0
        self.wrong_answers = []
        self.show_question()

    def save_errors(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if not file_path:
            return
        with open(file_path, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(["Вопрос", "Вариант A", "B", "C", "D", "E", "F", "Правильный ответ"])
            for q in self.wrong_answers:
                writer.writerow([q["question"]] + q["options"] + [q["answer"]])
        messagebox.showinfo("Сохранено", "Ошибки сохранены в файл.")

    def load_errors_csv(self):
        file_path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if not file_path:
            return
        self.test_questions = []
        try:
            with open(file_path, newline='', encoding='utf-8') as file:
                reader = csv.reader(file)
                next(reader)
                for row in reader:
                    self.test_questions.append({
                        "question": row[0],
                        "options": [opt if opt else "" for opt in row[1:7]],
                        "answer": row[7].strip().upper() if row[7] else ""
                    })
            if not self.test_questions:
                messagebox.showinfo("Пусто", "Нет доступных вопросов для теста.")
                return
            self.current_question = 0
            self.correct = 0
            self.incorrect = 0
            self.wrong_answers = []
            self.show_question()
        except Exception as e:
            messagebox.showerror("Ошибка загрузки CSV", str(e))

    def show_stats(self):
        self.clear_window()
        tk.Label(self.root, text="Статистика", font=("Arial", 24)).pack(pady=20)
        total_tests = self.stats.get("total_tests", 0)
        total_correct = self.stats.get("total_correct", 0)
        total_incorrect = self.stats.get("total_incorrect", 0)
        tk.Label(self.root, text=f"Всего тестов пройдено: {total_tests}", font=("Arial", 14)).pack(pady=5)
        tk.Label(self.root, text=f"Всего правильных ответов: {total_correct}", font=("Arial", 14)).pack(pady=5)
        tk.Label(self.root, text=f"Всего неправильных ответов: {total_incorrect}", font=("Arial", 14)).pack(pady=5)
        tk.Button(self.root, text="Назад в меню", command=self.main_menu, width=25, height=2).pack(pady=20)

    def save_stats(self, correct, incorrect):
        self.stats["total_tests"] = self.stats.get("total_tests", 0) + 1
        self.stats["total_correct"] = self.stats.get("total_correct", 0) + correct
        self.stats["total_incorrect"] = self.stats.get("total_incorrect", 0) + incorrect
        with open(self.stats_file, "w", encoding="utf-8") as f:
            json.dump(self.stats, f)

    def load_stats(self):
        if os.path.exists(self.stats_file):
            with open(self.stats_file, "r", encoding="utf-8") as f:
                return json.load(f)
        return {}

    def clear_window(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    def select_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.selected_file = file_path
            return True
        else:
            messagebox.showwarning("Предупреждение", "Файл не выбран. Тест не будет запущен.")
            return False

    def choose_test_mode(self):
        mode_window = tk.Toplevel(self.root)
        mode_window.title("Выбор режима")
        mode_window.geometry("300x150")
        tk.Label(mode_window, text="Выберите режим теста:", font=("Arial", 14)).pack(pady=10)

        tk.Button(mode_window, text="80 вопросов", width=25,
                  command=lambda: [mode_window.destroy(), self.handle_mode_selection(80)]).pack(pady=5)

        tk.Button(mode_window, text="Все вопросы (ХАРДКОР)", width=25,
                  command=lambda: [mode_window.destroy(), self.handle_mode_selection("all")]).pack(pady=5)

    def handle_mode_selection(self, mode):
        if not self.select_file():
            return
        try:
            wb = load_workbook(self.selected_file)
            sheet = wb.active
            self.questions = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is None for cell in row):
                    continue
                self.questions.append({
                    "question": row[1],
                    "options": [opt if opt is not None else "" for opt in row[2:8]],
                    "answer": row[8].strip().upper() if row[8] else ""
                })

            if len(self.questions) < 1:
                messagebox.showerror("Ошибка", "В файле нет вопросов.")
                return

            if mode == 80:
                if len(self.questions) < 80:
                    messagebox.showerror("Ошибка", "В файле меньше 80 вопросов.")
                    return
                self.questions = random.sample(self.questions, 80)

            self.test_questions = self.questions
            self.current_question = 0
            self.correct = 0
            self.incorrect = 0
            self.wrong_answers = []
            self.show_question()

        except Exception as e:
            messagebox.showerror("Ошибка загрузки", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    app = TestApp(root)
    root.mainloop()
